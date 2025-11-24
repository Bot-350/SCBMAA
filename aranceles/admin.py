from django.contrib import admin
from django.http import HttpResponse
from .models import Seccion, Capitulo, Partida, Subpartida, Nota, ItemNota, LogActualizacion

from import_export.admin import ImportExportModelAdmin
from import_export import resources, fields
from import_export.widgets import ForeignKeyWidget, DecimalWidget, BooleanWidget
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, PatternFill
import uuid
from datetime import datetime

# -----------------------------
# WIDGET DECIMAL PERSONALIZADO
# -----------------------------
class CustomDecimalWidget(DecimalWidget):
    def clean(self, value, row=None, *args, **kwargs):
        if value in [None, '']:
            return None
        try:
            return float(str(value).replace(',', '.'))
        except (ValueError, TypeError):
            return None

# -----------------------------
# FUNCION DE LOG COMUN
# -----------------------------
def registrar_log(modelo, instance, action, detalles=""):
    LogActualizacion.objects.create(
        fecha_inicio=datetime.now(),
        status=action,
        filas_agregadas=1 if action == "Creado" else 0,
        filas_modificadas=1 if action == "Modificado" else 0,
        filas_eliminadas=1 if action == "Eliminado" else 0,
        detalles=f"{modelo}: {getattr(instance, 'codigo', getattr(instance, 'nombre', str(instance)))} {detalles}"
    )

# -----------------------------
# ADMIN BASE CON LOG
# -----------------------------
class BaseLoggedAdmin(admin.ModelAdmin):
    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        action = "Modificado" if change else "Creado"
        registrar_log(obj.__class__.__name__, obj, action)

    def delete_model(self, request, obj):
        registrar_log(obj.__class__.__name__, obj, "Eliminado")
        super().delete_model(request, obj)

# -----------------------------
# ADMIN DE EXPORTACIÓN ESTILIZADA PARA SUBPARTIDAS
# -----------------------------
class StyledExportAdmin(ImportExportModelAdmin, BaseLoggedAdmin):
    def export_action(self, request, *args, **kwargs):
        queryset = self.get_export_queryset(request)
        # Ordenar por el campo 'orden' para respetar la posición de títulos
        queryset = queryset.order_by('orden')
        
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = 'Tabla Aranceles'
        # Headers principales
        headers = [
            'CÓDIGO', 'CÓDIGO PARTIDA', 'DESCRIPCIÓN', 'GA %', 'ICE/IEHD', 'UNIDAD DE MEDIDA', 'DESPACHO EN FRONTERA',
            'TIPO DOC', 'ENTIDAD QUE EMITE', 'DISP. LEGAL',
            'CAN ACE 36 ACE 47 VEN', 'ACE 22 Chi', 'ACE 22 Prot', 'ACE 66 MEXICO', 'ES TÍTULO'
        ]
        sheet.append(headers)

        # Colores azulados
        header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")  # Azul fuerte
        alt_fill_1 = PatternFill(start_color="F5F8FD", end_color="F5F8FD", fill_type="solid")  # Azul muy suave
        alt_fill_2 = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")  # Azul claro
        titulo_fill = PatternFill(start_color="90CAF9", end_color="90CAF9", fill_type="solid")  # Azul destacado

        # Encabezados con color y negrita
        for cell in sheet[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill

        # Exportar en el orden exacto del queryset (ordenado por 'orden')
        for idx, obj in enumerate(queryset, start=2):
            # Marcar como título intermedio si el código empieza por '_H_'
            es_titulo = obj.codigo.startswith('_H_') or obj.es_titulo_intermedio
            es_titulo_str = 'Sí' if es_titulo else 'No'
            row = [
                obj.codigo or "",
                getattr(obj.partida, 'codigo', "") or "",
                obj.descripcion or "",
                obj.ga if obj.ga is not None else "",
                getattr(obj, 'ice_iehd', "") or "",
                getattr(obj, 'unidad_medida', "") or "",
                getattr(obj, 'despacho_frontera', "") or "",
                getattr(obj, 'tipo_de_doc', "") or "",
                getattr(obj, 'entidad_que_emite', "") or "",
                getattr(obj, 'disposicion_legal', "") or "",
                getattr(obj, 'can_ace_36_47_ven', '') or "",
                getattr(obj, 'ace_22_chile', '') or "",
                getattr(obj, 'ace_22_prot', '') or "",
                getattr(obj, 'ace_66_mexico', '') or "",
                es_titulo_str
            ]
            sheet.append(row)
            # Alternar colores de fondo en filas
            fill = alt_fill_1 if idx % 2 == 0 else alt_fill_2
            if es_titulo:
                fill = titulo_fill
            for cell in sheet[idx]:
                cell.fill = fill
            if es_titulo:
                for cell in sheet[idx]:
                    cell.font = Font(bold=True, color="0D47A1")

        # Ajustar ancho de columnas
        for i, column_cells in enumerate(sheet.columns):
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            column_letter = get_column_letter(i + 1)
            sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

        # Agregar filtros automáticos a todas las columnas
        sheet.auto_filter.ref = sheet.dimensions

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="Tabla_Aranceles.xlsx"'
        workbook.save(response)
        return response
    export_action.short_description = "Exportar seleccionados a Excel"

# -----------------------------
# RECURSO PARA SUBPARTIDAS
# -----------------------------
class SubpartidaOrdenadaResource(resources.ModelResource):
    codigo = fields.Field(
        attribute='codigo',
        column_name='CÓDIGO'
    )
    partida = fields.Field(
        column_name='CÓDIGO PARTIDA',
        attribute='partida',
        widget=ForeignKeyWidget(Partida, 'codigo')
    )
    descripcion = fields.Field(
        attribute='descripcion',
        column_name='DESCRIPCIÓN'
    )
    ga = fields.Field(
        attribute='ga',
        column_name='GA %',
        widget=CustomDecimalWidget()
    )
    unidad_medida = fields.Field(
        attribute='unidad_medida',
        column_name='UNIDAD DE MEDIDA'
    )
    tipo_de_doc = fields.Field(
        attribute='tipo_de_doc',
        column_name='TIPO DOC'
    )
    entidad_que_emite = fields.Field(
        attribute='entidad_que_emite',
        column_name='ENTIDAD QUE EMITE'
    )
    disposicion_legal = fields.Field(
        attribute='disposicion_legal',
        column_name='DISP. LEGAL'
    )
    can_ace_36_47_ven = fields.Field(
        attribute='can_ace_36_47_ven',
        column_name='CAN ACE 36 ACE 47 VEN'
    )
    ace_22_chile = fields.Field(
        attribute='ace_22_chile',
        column_name='ACE 22 Chi'
    )
    ace_22_prot = fields.Field(
        attribute='ace_22_prot',
        column_name='ACE 22 Prot'
    )
    ace_66_mexico = fields.Field(
        attribute='ace_66_mexico',
        column_name='ACE 66 MEXICO'
    )
    es_titulo_intermedio = fields.Field(
        attribute='es_titulo_intermedio',
        column_name='ES TÍTULO',
        widget=BooleanWidget()
    )

    class Meta:
        model = Subpartida
        import_id_fields = ('codigo',)
        fields = (
            'codigo', 'partida', 'descripcion', 'ga', 'unidad_medida',
            'tipo_de_doc', 'entidad_que_emite', 'disposicion_legal',
            'can_ace_36_47_ven', 'ace_22_chile', 'ace_22_prot', 'ace_66_mexico',
            'es_titulo_intermedio', 'orden'
        )
        skip_unchanged = True
        report_skipped = True

    def before_import(self, dataset, **kwargs):
        self.codigos_antes = set(Subpartida.objects.values_list('codigo', flat=True))
        super().before_import(dataset, **kwargs)

    def before_import_row(self, row, **kwargs):
        # Normalizar los nombres de columna: convertir a mayúsculas para reconocer variaciones
        row_normalized = {}
        for key, value in row.items():
            # Mapeo de posibles variaciones de nombres de columna
            key_upper = str(key).strip().upper()
            # Limpiar espacios en blanco de los valores
            if isinstance(value, str):
                value = value.strip()
            row_normalized[key_upper] = value
        
        # Si es un título intermedio sin código, generar uno automáticamente
        es_titulo_valores = ['1', 1, 'true', 'True', True, 'Sí', 'SI', 'si', 'SÍ']
        es_titulo = row_normalized.get('ES TÍTULO')
        codigo_actual = row_normalized.get('CÓDIGO')
        if es_titulo in es_titulo_valores and (not codigo_actual or str(codigo_actual).strip() == ''):
            row_normalized['CÓDIGO'] = f"_H_{uuid.uuid4().hex[:10].upper()}"
        
        # Asegurar que siempre hay un CÓDIGO PARTIDA válido (después de limpiar espacios)
        codigo_partida = row_normalized.get('CÓDIGO PARTIDA')
        if not codigo_partida or (isinstance(codigo_partida, str) and not codigo_partida.strip()):
            raise ValueError(f"CÓDIGO PARTIDA es requerido - fila tiene: '{codigo_partida}'")
        
        # Actualizar la row con los valores normalizados
        row.clear()
        row.update(row_normalized)
        super().before_import_row(row, **kwargs)

    def after_import(self, dataset, result, **kwargs):
        from aranceles.models import RegistroCambio, Subpartida
        dry_run = kwargs.get('dry_run', False)
        if dry_run:
            return

        nuevas = sum(1 for row in result.rows if row.import_type == row.IMPORT_TYPE_NEW)
        modificadas = sum(1 for row in result.rows if row.import_type == row.IMPORT_TYPE_UPDATE)
        omitidas = sum(1 for row in result.rows if row.import_type == row.IMPORT_TYPE_SKIP)

        codigos_importados = set()
        for row in dataset.dict:
            codigo = row.get('CÓDIGO')
            if codigo:
                codigos_importados.add(str(codigo).strip())

        codigos_antes = getattr(self, 'codigos_antes', set())
        codigos_a_eliminar = codigos_antes - codigos_importados
        eliminadas = 0
        if codigos_a_eliminar:
            eliminadas = len(codigos_a_eliminar)
            subpartidas_eliminadas = list(Subpartida.objects.filter(codigo__in=codigos_a_eliminar))
            Subpartida.objects.filter(codigo__in=codigos_a_eliminar).delete()
            # Registrar eliminaciones
            for sp in subpartidas_eliminadas:
                RegistroCambio.objects.create(
                    tipo_cambio='eliminar',
                    modelo='Subpartida',
                    objeto_id=sp.id,
                    usuario='Sistema',
                    descripcion=f"Eliminado por importación: {sp.codigo}",
                    cambios_detalles={}
                )

        # Registrar nuevas y modificadas
        for row, import_row in zip(dataset.dict, result.rows):
            codigo = row.get('CÓDIGO')
            if not codigo:
                continue
            try:
                sp = Subpartida.objects.get(codigo=str(codigo).strip())
            except Subpartida.DoesNotExist:
                continue
            if import_row.import_type == import_row.IMPORT_TYPE_NEW:
                RegistroCambio.objects.create(
                    tipo_cambio='crear',
                    modelo='Subpartida',
                    objeto_id=sp.id,
                    usuario='Sistema',
                    descripcion=f"Creado por importación: {sp.codigo}",
                    cambios_detalles={}
                )
            elif import_row.import_type == import_row.IMPORT_TYPE_UPDATE:
                RegistroCambio.objects.create(
                    tipo_cambio='actualizar',
                    modelo='Subpartida',
                    objeto_id=sp.id,
                    usuario='Sistema',
                    descripcion=f"Actualizado por importación: {sp.codigo}",
                    cambios_detalles={}
                )

        LogActualizacion.objects.create(
            fecha_inicio=datetime.now(),
            status='Completado',
            filas_agregadas=nuevas,
            filas_modificadas=modificadas,
            filas_eliminadas=eliminadas,
            detalles=f"Importación finalizada: {nuevas} nuevas, {modificadas} modificadas, {eliminadas} eliminadas, {omitidas} omitidas."
        )

        # Mostrar mensaje de confirmación con el número correcto de eliminados
        from import_export.results import RowResult
        result.deleted_rows = eliminadas  # Forzar el valor correcto en el resultado

        super().after_import(dataset, result, **kwargs)

# -----------------------------
# ADMIN DE MODELOS
# -----------------------------
@admin.register(Seccion)
class SeccionAdmin(BaseLoggedAdmin):
    list_display = ('id', 'nombre', 'descripcion')
    ordering = ('id',)

@admin.register(Capitulo)
class CapituloAdmin(BaseLoggedAdmin):
    list_display = ('codigo', 'nombre', 'seccion')
    list_filter = ('seccion',)
    ordering = ('codigo',)

@admin.register(Partida)
class PartidaAdmin(BaseLoggedAdmin):
    list_display = ('codigo', 'descripcion', 'capitulo')
    list_filter = ('capitulo__seccion', 'capitulo')
    search_fields = ('codigo', 'descripcion')
    ordering = ('codigo',)

class ItemNotaInline(admin.TabularInline):
    model = ItemNota
    extra = 1

    def save_model(self, request, obj, form, change):
        super().save_model(request, obj, form, change)
        action = "Modificado" if change else "Creado"
        registrar_log(obj.__class__.__name__, obj, action)

    def delete_model(self, request, obj):
        registrar_log(obj.__class__.__name__, obj, "Eliminado")
        super().delete_model(request, obj)

@admin.register(Nota)
class NotaAdmin(BaseLoggedAdmin):
    list_display = ('id', 'tipo', 'texto_corto', 'seccion', 'capitulo')
    list_filter = ('tipo', 'seccion', 'capitulo')
    inlines = [ItemNotaInline]

    def texto_corto(self, obj):
        if getattr(obj, 'es_lista', False):
            return getattr(obj, 'titulo_lista', '')
        return (getattr(obj, 'texto', '')[:75] + '...') if len(getattr(obj, 'texto', '')) > 75 else getattr(obj, 'texto', '')

@admin.register(Subpartida)
class SubpartidaAdmin(StyledExportAdmin):
    resource_class = SubpartidaOrdenadaResource
    actions = ['export_action']
    list_display = ('orden', 'codigo', 'descripcion', 'ga', 'unidad_medida', 'partida', 'es_titulo_intermedio')
    list_filter = ('partida__capitulo__seccion', 'partida__capitulo', 'es_titulo_intermedio')
    search_fields = ('codigo', 'descripcion')
    ordering = ('orden',)

@admin.register(LogActualizacion)
class LogActualizacionAdmin(admin.ModelAdmin):
    list_display = ('fecha_inicio', 'status', 'filas_agregadas', 'filas_modificadas', 'filas_eliminadas', 'detalles')
    ordering = ('-fecha_inicio',)
